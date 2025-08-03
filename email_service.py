import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from datetime import date, datetime
import pytz
from app import db
import os
import tempfile
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch

# Indian Standard Time
IST = pytz.timezone('Asia/Kolkata')
from models import User, Attendance, EmailLog
import logging
import os

# Email configuration
SMTP_SERVER = "smtp.gmail.com"
SMTP_PORT = 587
SMTP_EMAIL = os.getenv("SMTP_EMAIL", "marcus189076@gmail.com")
SMTP_PASSWORD = os.getenv("SMTP_PASSWORD", "ggeg knjd qtku ssrb")

ADMIN_EMAILS = [
    "nlramcharanplacement@gmail.com",
    "2224m1a3133@vemu.org"
]

def generate_excel_report(present_faculty, absent_faculty, today_attendances, today):
    """Generate Excel report for attendance"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Report"
    
    # Title
    ws['A1'] = f"Daily Attendance Report - {today.strftime('%B %d, %Y')}"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A1:E1')
    
    # Headers
    headers = ['S.No', 'Faculty Name', 'Username', 'Department', 'Status', 'Scan Time']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    row = 4
    
    # Present faculty
    for i, faculty in enumerate(present_faculty, 1):
        attendance = next((att for att, user in today_attendances if user.id == faculty.id), None)
        if attendance:
            ist_time = attendance.scan_time.replace(tzinfo=pytz.UTC).astimezone(IST)
            scan_time = ist_time.strftime('%I:%M %p IST')
        else:
            scan_time = 'N/A'
            
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=faculty.full_name or faculty.username)
        ws.cell(row=row, column=3, value=faculty.username)
        ws.cell(row=row, column=4, value=faculty.department or 'N/A')
        ws.cell(row=row, column=5, value='Present')
        ws.cell(row=row, column=6, value=scan_time)
        
        # Green background for present
        for col in range(1, 7):
            ws.cell(row=row, column=col).fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")
        
        row += 1
    
    # Absent faculty
    for i, faculty in enumerate(absent_faculty, len(present_faculty) + 1):
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=faculty.full_name or faculty.username)
        ws.cell(row=row, column=3, value=faculty.username)
        ws.cell(row=row, column=4, value=faculty.department or 'N/A')
        ws.cell(row=row, column=5, value='Absent')
        ws.cell(row=row, column=6, value='N/A')
        
        # Red background for absent
        for col in range(1, 7):
            ws.cell(row=row, column=col).fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
        
        row += 1
    
    # Summary
    row += 2
    ws.cell(row=row, column=1, value='Summary').font = Font(bold=True)
    row += 1
    ws.cell(row=row, column=1, value=f'Total Faculty: {len(present_faculty) + len(absent_faculty)}')
    row += 1
    ws.cell(row=row, column=1, value=f'Present: {len(present_faculty)}')
    row += 1
    ws.cell(row=row, column=1, value=f'Absent: {len(absent_faculty)}')
    row += 1
    attendance_rate = (len(present_faculty) / (len(present_faculty) + len(absent_faculty)) * 100) if (len(present_faculty) + len(absent_faculty)) > 0 else 0
    ws.cell(row=row, column=1, value=f'Attendance Rate: {attendance_rate:.1f}%')
    
    # Auto-adjust column widths
    column_letters = ['A', 'B', 'C', 'D', 'E', 'F']
    column_widths = [8, 25, 15, 20, 12, 20]  # Default widths for each column
    
    for i, width in enumerate(column_widths):
        if i < len(column_letters):
            ws.column_dimensions[column_letters[i]].width = width
    
    # Save to temporary file
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(temp_file.name)
    return temp_file.name

def generate_pdf_report(present_faculty, absent_faculty, today_attendances, today):
    """Generate PDF report for attendance"""
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
    doc = SimpleDocTemplate(temp_file.name, pagesize=A4)
    story = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        alignment=1,  # Center
        spaceAfter=30
    )
    
    # Title
    title = Paragraph(f"Daily Attendance Report - {today.strftime('%B %d, %Y')}", title_style)
    story.append(title)
    
    # Present Faculty Table
    if present_faculty:
        story.append(Paragraph("Present Faculty", styles['Heading2']))
        present_data = [['S.No', 'Faculty Name', 'Username', 'Department', 'Scan Time']]
        
        for i, faculty in enumerate(present_faculty, 1):
            attendance = next((att for att, user in today_attendances if user.id == faculty.id), None)
            if attendance:
                ist_time = attendance.scan_time.replace(tzinfo=pytz.UTC).astimezone(IST)
                scan_time = ist_time.strftime('%I:%M %p IST')
            else:
                scan_time = 'N/A'
                
            present_data.append([
                str(i),
                faculty.full_name or faculty.username,
                faculty.username,
                faculty.department or 'N/A',
                scan_time
            ])
        
        present_table = Table(present_data)
        present_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.lightgreen),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(present_table)
        story.append(Spacer(1, 20))
    
    # Absent Faculty Table
    if absent_faculty:
        story.append(Paragraph("Absent Faculty", styles['Heading2']))
        absent_data = [['S.No', 'Faculty Name', 'Username', 'Department']]
        
        for i, faculty in enumerate(absent_faculty, 1):
            absent_data.append([
                str(i),
                faculty.full_name or faculty.username,
                faculty.username,
                faculty.department or 'N/A'
            ])
        
        absent_table = Table(absent_data)
        absent_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.lightcoral),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(absent_table)
        story.append(Spacer(1, 20))
    
    # Summary
    story.append(Paragraph("Summary", styles['Heading2']))
    total_faculty = len(present_faculty) + len(absent_faculty)
    attendance_rate = (len(present_faculty) / total_faculty * 100) if total_faculty > 0 else 0
    
    summary_data = [
        ['Total Faculty', str(total_faculty)],
        ['Present', str(len(present_faculty))],
        ['Absent', str(len(absent_faculty))],
        ['Attendance Rate', f'{attendance_rate:.1f}%']
    ]
    
    summary_table = Table(summary_data)
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.lightblue),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    story.append(summary_table)
    
    # Build PDF
    doc.build(story)
    return temp_file.name

def send_attendance_email():
    """Send attendance report email to administrators"""
    try:
        # Get today's date
        today = date.today()
        
        # Get all faculty members
        all_faculty = User.query.filter_by(role='faculty').all()
        
        # Get today's attendances
        today_attendances = db.session.query(Attendance, User).join(User).filter(
            Attendance.date == today
        ).all()
        
        # Create lists of present and absent faculty
        present_faculty = [user for att, user in today_attendances]
        present_faculty_ids = [f.id for f in present_faculty]
        absent_faculty = [f for f in all_faculty if f.id not in present_faculty_ids]
        
        # Create email content
        subject = f"Daily Attendance Report - {today.strftime('%B %d, %Y')}"
        
        html_body = f"""
        <html>
        <body>
            <h2>Daily Attendance Report</h2>
            <p><strong>Date:</strong> {today.strftime('%B %d, %Y')}</p>
            
            <h3>Present Faculty ({len(present_faculty)})</h3>
            <ul>
        """
        
        for faculty in present_faculty:
            attendance = next((att for att, user in today_attendances if user.id == faculty.id), None)
            if attendance:
                # Convert to IST for display
                ist_time = attendance.scan_time.replace(tzinfo=pytz.UTC).astimezone(IST)
                scan_time = ist_time.strftime('%I:%M %p IST')
            else:
                scan_time = 'N/A'
            html_body += f"<li>{faculty.full_name} ({faculty.username}) - {faculty.department} - Scanned at: {scan_time}</li>"
        
        html_body += """
            </ul>
            
            <h3>Absent Faculty ({})
            <ul>
        """.format(len(absent_faculty))
        
        for faculty in absent_faculty:
            html_body += f"<li>{faculty.full_name} ({faculty.username}) - {faculty.department}</li>"
        
        html_body += """
            </ul>
            
            <p><strong>Total Faculty:</strong> {}</p>
            <p><strong>Present:</strong> {}</p>
            <p><strong>Absent:</strong> {}</p>
            <p><strong>Attendance Rate:</strong> {:.1f}%</p>
            
            <hr>
            <p><em>This is an automated email from the Smart Attendance System.</em></p>
        </body>
        </html>
        """.format(
            len(all_faculty),
            len(present_faculty),
            len(absent_faculty),
            (len(present_faculty) / len(all_faculty) * 100) if all_faculty else 0
        )
        
        # Generate Excel and PDF reports
        excel_file = generate_excel_report(present_faculty, absent_faculty, today_attendances, today)
        pdf_file = generate_pdf_report(present_faculty, absent_faculty, today_attendances, today)
        
        # Prepare attachments
        attachments = [
            {
                'filename': f'attendance_report_{today.strftime("%Y%m%d")}.xlsx',
                'filepath': excel_file
            },
            {
                'filename': f'attendance_report_{today.strftime("%Y%m%d")}.pdf',
                'filepath': pdf_file
            }
        ]
        
        # Send email with attachments
        success = send_email_with_attachments(ADMIN_EMAILS, subject, html_body, attachments)
        
        # Clean up temporary files
        try:
            os.unlink(excel_file)
            os.unlink(pdf_file)
        except:
            pass
        
        # Log email
        email_log = EmailLog(
            date=today,
            recipients=', '.join(ADMIN_EMAILS),
            subject=subject,
            status='sent' if success else 'failed'
        )
        db.session.add(email_log)
        db.session.commit()
        
        return success
        
    except Exception as e:
        logging.error(f"Error sending attendance email: {e}")
        return False

def send_email(recipients, subject, html_body):
    """Send email using SMTP"""
    try:
        # Create message
        msg = MIMEMultipart('alternative')
        msg['Subject'] = subject
        msg['From'] = SMTP_EMAIL
        msg['To'] = ', '.join(recipients)
        
        # Add HTML content
        html_part = MIMEText(html_body, 'html')
        msg.attach(html_part)
        
        # Send email
        server = smtplib.SMTP(SMTP_SERVER, SMTP_PORT)
        server.starttls()
        server.login(SMTP_EMAIL, SMTP_PASSWORD)
        
        for recipient in recipients:
            server.sendmail(SMTP_EMAIL, recipient, msg.as_string())
        
        server.quit()
        logging.info(f"Email sent successfully to {recipients}")
        return True
        
    except Exception as e:
        logging.error(f"Failed to send email: {e}")
        return False

def send_email_with_attachments(recipients, subject, html_body, attachments):
    """Send email with attachments using SMTP"""
    try:
        # Create message
        msg = MIMEMultipart()
        msg['Subject'] = subject
        msg['From'] = SMTP_EMAIL
        msg['To'] = ', '.join(recipients)
        
        # Add HTML content
        html_part = MIMEText(html_body, 'html')
        msg.attach(html_part)
        
        # Add attachments
        for attachment in attachments:
            with open(attachment['filepath'], 'rb') as file:
                part = MIMEBase('application', 'octet-stream')
                part.set_payload(file.read())
                encoders.encode_base64(part)
                part.add_header(
                    'Content-Disposition',
                    f'attachment; filename= {attachment["filename"]}'
                )
                msg.attach(part)
        
        # Send email
        server = smtplib.SMTP(SMTP_SERVER, SMTP_PORT)
        server.starttls()
        server.login(SMTP_EMAIL, SMTP_PASSWORD)
        
        for recipient in recipients:
            server.sendmail(SMTP_EMAIL, recipient, msg.as_string())
        
        server.quit()
        logging.info(f"Email with attachments sent successfully to {recipients}")
        return True
        
    except Exception as e:
        logging.error(f"Failed to send email with attachments: {e}")
        return False
